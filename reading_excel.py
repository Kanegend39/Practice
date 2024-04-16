import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from scipy.signal import hilbert
import math

path = "50.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
start = sheet_obj.cell(row=2, column=1).value
x = np.arange(start, start + m_row - 1, 1)
grafik_1 = np.arange(0, m_row - 1, 1).astype("float")
grafik_2 = np.arange(0, m_row - 1, 1).astype("float")
grafik_3 = np.arange(0, m_row - 1, 1).astype("float")
grafik_4 = np.arange(0, m_row - 1, 1).astype("float")
print(x)
print(m_row)
for i in range(2, m_row + 1):
    obj_1 = sheet_obj.cell(row=i, column=2)
    grafik_1[i - 2] = obj_1.value
    obj_2 = sheet_obj.cell(row=i, column=4)
    grafik_2[i - 2] = obj_2.value
    obj_3 = sheet_obj.cell(row=i, column=6)
    grafik_3[i - 2] = obj_3.value
    obj_4 = sheet_obj.cell(row=i, column=8)
    grafik_4[i - 2] = obj_4.value
#print(grafik_1)
#print(grafik_2)
print(grafik_3)
print(grafik_4)
I_3 = 0
I_3 = hilbert(grafik_3)  # преобразование Гильберта
a3 = np.sqrt(I_3.real ** 2 + I_3.imag ** 2)  # амплитуда
comp3 = np.arctan2(I_3.imag, I_3.real)
approximated_line_3 = np.polyfit(x, np.unwrap(comp3, math.pi), 1)
I_4 = 0
I_4 = hilbert(grafik_4)  # преобразование Гильберта
a4 = np.sqrt(I_4.real ** 2 + I_4.imag ** 2)  # амплитуда
comp4 = np.arctan2(I_4.imag, I_4.real)
approximated_line_4 = np.polyfit(x, np.unwrap(comp4, math.pi), 1)
plt.rcParams.update({
    'axes.facecolor': 'black'
})
phi_3 = np.unwrap(comp3, math.pi) - (x * approximated_line_3[0] + approximated_line_3[1])
phi_4 = np.unwrap(comp4, math.pi) - (x * approximated_line_4[0] + approximated_line_4[1])
#plt.plot(x, grafik_1, color='white')
#plt.plot(x, grafik_2, color='red')
plt.plot(x, grafik_3, color='green')
plt.plot(x, grafik_4, color='blue')
plt.plot(x, phi_4 - phi_3, color='violet')
plt.show()
